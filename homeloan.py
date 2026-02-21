import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# --- CONFIGURATION ---
# Target Directory requested
output_dir = "/home/andrew/.ssh/Trading/_MelbInvestments/output"
file_name = "Investment_Portfolio_Tracker.xlsx"
full_path = os.path.join(output_dir, file_name)

# Ensure directory exists
if not os.path.exists(output_dir):
    try:
        os.makedirs(output_dir, exist_ok=True)
        print(f"Created directory: {output_dir}")
    except PermissionError:
        print(f"Error: Permission denied creating {output_dir}. Check your .ssh folder permissions.")
        exit()

# --- CREATE WORKBOOK ---
wb = openpyxl.Workbook()

# --- STYLES ---
header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill("solid", fgColor="2F75B5")  # Blue
currency_format = '"$"#,##0.00'
percent_format = '0.00%'
center_align = Alignment(horizontal='center', vertical='center')

def style_range(ws, cell_range, fill=None, font=None, alignment=None):
    rows = ws[cell_range]
    for row in rows:
        for cell in row:
            if fill: cell.fill = fill
            if font: cell.font = font
            if alignment: cell.alignment = alignment

# --- SHEET 1: DASHBOARD ---
ws1 = wb.active
ws1.title = "Portfolio Dashboard"
ws1['B2'] = "PORTFOLIO SUMMARY"
ws1['B2'].font = Font(bold=True, size=14)

# Dashboard Formulas
labels = {
    'B4': "Total Property Value", 'C4': "=SUM('Property Details'!C4:C5)",
    'B5': "Total Debt Position",  'C5': "=SUM('Loan Manager'!C4:C6)",
    'B6': "Net Equity",           'C6': "=C4-C5",
    'B7': "Overall LVR",          'C7': "=C5/C4",
    'E4': "Monthly Rental Income",'F4': "='Cash Flow'!C4",
    'E5': "Monthly Mortgage Cost",'F5': "=SUM('Loan Manager'!F4:F6)",
    'E6': "Monthly OpEx (Est.)",  'F6': "='Cash Flow'!C14",
    'E7': "Net Monthly Cashflow", 'F7': "=F4-(F5+F6)"
}
for cell, val in labels.items():
    ws1[cell] = val

for c in ['C4','C5','C6','F4','F5','F6','F7']: ws1[c].number_format = currency_format
ws1['C7'].number_format = percent_format

# --- SHEET 2: LOAN MANAGER ---
ws2 = wb.create_sheet("Loan Manager")
ws2.append(["Loan ID", "Purpose", "Loan Amount", "Interest Rate", "Type", "Monthly Pmt", "IO Expiry", "Lender"])

loans = [
    ["Loan 1.1", "Home Refinance (OO)", 445826, 0.0525, "P&I", 2461.87, "N/A", "Mortgage Choice"],
    ["Loan 1.2", "Inv Deposit (Cash Out)", 170000, 0.0549, "P&I", 964.17, "N/A", "Mortgage Choice"],
    ["Loan 2.1", "Investment Purchase", 520000, 0.0564, "IO (5yr)", 2444.00, "Jan-2031", "Mortgage Choice"]
]
for row in loans: ws2.append(row)

# Totals
ws2['B6'] = "TOTALS"; ws2['B6'].font = Font(bold=True)
ws2['C6'] = "=SUM(C2:C5)"; ws2['C6'].number_format = currency_format
ws2['F6'] = "=SUM(F2:F5)"; ws2['F6'].number_format = currency_format

style_range(ws2, "A1:H1", fill=header_fill, font=header_font, alignment=center_align)
for row in range(2, 6):
    ws2[f'C{row}'].number_format = currency_format
    ws2[f'D{row}'].number_format = percent_format
    ws2[f'F{row}'].number_format = currency_format

# --- SHEET 3: PROPERTY DETAILS ---
ws3 = wb.create_sheet("Property Details")
ws3.append(["Property", "Purchase Price", "Current Value", "Purchase Date", "Stamp Duty"])
props = [
    ["Home (Owner Occ)", 1200000, 1200000, "Existing", 0],
    ["Investment 1", 650000, 650000, "Jan-2026", 34070]
]
for row in props: ws3.append(row)

style_range(ws3, "A1:E1", fill=header_fill, font=header_font)
for col in ['B', 'C', 'E']:
    for r in range(2,4): ws3[f'{col}{r}'].number_format = currency_format

# --- SHEET 4: CASH FLOW ---
ws4 = wb.create_sheet("Cash Flow")
ws4.column_dimensions['B'].width = 25
ws4['B2'] = "INCOME"; ws4['B2'].font = Font(bold=True)
ws4['B3'] = "Weekly Rent (Est.)"; ws4['C3'] = 650
ws4['B4'] = "Monthly Rent"; ws4['C4'] = "=(C3*52)/12"

ws4['B6'] = "EXPENSES (Monthly)"; ws4['B6'].font = Font(bold=True)
expenses = [
    ("Council Rates", 200), ("Water Rates", 100), ("Body Corp", 0),
    ("Mgmt Fees (8%)", "=C4*0.08"), ("Insurance", 150), ("Maintenance", 100)
]
r = 7
for name, val in expenses:
    ws4[f'B{r}'] = name; ws4[f'C{r}'] = val
    r += 1
ws4[f'B{r}'] = "TOTAL OPEX"; ws4[f'C{r}'] = f"=SUM(C7:C{r-1})"

# --- SAVE FILE ---
try:
    wb.save(full_path)
    print(f"Success! File saved to: {full_path}")
except Exception as e:
    print(f"Error saving file: {e}")